from openpyxl import Workbook,load_workbook

wb = load_workbook("C:\\Users\\ataka\\Desktop\\Sitev\\veri.xlsx")
ws = wb.active
print(wb.sheetnames) 
for satir in range(1,9):
    for sutun in range(1,9):        
        print(" | " + str(ws.cell(satir,sutun).value) + " | ",end="")
    print()