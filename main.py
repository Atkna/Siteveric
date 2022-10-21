from asyncore import write
from lib2to3.pytree import convert
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook,load_workbook
import csv
import numpy as np
import xlsxwriter
import xlrd
import matplotlib.pyplot as plt
import seaborn as sns



eksen = open("C:\\Users\\3drea\\Desktop\\Siteveric-main\\3eksen.txt")

eksenoku = eksen.read()

url = eksenoku

dream = open("C:\\Users\\3drea\\Desktop\\Siteveric-main\\3dream.txt")

dreamoku = dream.read()

url2 = dreamoku

robo = open("C:\\Users\\3drea\\Desktop\\Siteveric-main\\robo.txt")

robooku = robo.read()

url3 = robooku

robon11 = open("C:\\Users\\3drea\\Desktop\\Siteveric-main\\robon11.txt")

robon11oku = robon11.read()

url4 = robon11oku

#dolar = open("C:\\Users\\3drea\\Desktop\\Siteveric-main\\dolar.txt")

#dollar = dolar.read()

urldolar = "https://www.google.com/finance/quote/USD-TRY"

#dolar
response = requests.get(urldolar)
html_icerigi = response.content
soup = BeautifulSoup(html_icerigi,"html.parser")

dolarfiyat = soup.find("div",{"class":"YMlKec fxKbKc"})

dolarfiyat = (dolarfiyat.text).strip("\n").strip("\t").replace(".",",")

dolarson = dolarfiyat

dolarson2 = dolarson 

print(dolarson)
#3eksen
response = requests.get(url)
html_icerigi = response.content
soup = BeautifulSoup(html_icerigi,"html.parser")

eskifiyat = soup.find_all("div",{"class":"showcase-price-old"})
fiyat = soup.find_all("div",{"class":"showcase-price-new"})
isim = soup.find_all("div",{"class":"showcase-title"})


liste = list()

for i in range(len(isim)):
    
   
    mEskiFiyat = ""
    isim[i] = (isim[i].text).strip("\n").strip("\t")
    fiyat[i] = (fiyat[i].text).strip("\n").replace("TL","").strip()
    try:
     eskifiyat[i] = (eskifiyat[i].text).strip("\n").replace("\nTL","").strip() 
     mEskiFiyat = eskifiyat[i]
    except IndexError:
        mEskiFiyat = ""
    liste.append([isim[i],fiyat[i],dolarson])#,mEskiFiyat


#3dream
response = requests.get(url2)
html_icerigi = response.content
soup = BeautifulSoup(html_icerigi,"html.parser")

reskifiyat = soup.find_all("class",{"ins":"price dib mb__5"})
rfiyat = soup.find_all("span",{"class":"price dib mb__5"})

if isinstance(rfiyat, type(list)):
  rfiyat = rfiyat[1]
risim = soup.find_all("a",{"class":"cd chp"})

rliste = list()

for i in range(len(risim)):
 
    rmEskiFiyat = ""
    risim[i] = (risim[i].text).strip("\n").strip("\n")
    rfiyat[i] = (rfiyat[i].text).strip("\n").replace("TL"," TL").strip("\n")
    try:
     reskifiyat[i] = (reskifiyat[i].text).strip("").replace("TL"," TL").strip() 
     rmEskiFiyat = reskifiyat[i]
    except IndexError:
        rmEskiFiyat = ""
    rliste.append([risim[i],rfiyat[i]])
    
     #roboboloq

     
response = requests.get(url3)
html_icerigi = response.content
soup = BeautifulSoup(html_icerigi,"html.parser")

roboeskifiyat = soup.find_all("div",{"class":"showcase-price-old"})
robofiyat = soup.find_all("div",{"class":"discountPrice"})
roboisim = soup.find_all("div",{"class":"productName detailUrl"})

roboliste = list()

for i in range(len(roboisim)):
 
    robomEskiFiyat = ""
    roboisim[i] = (roboisim[i].text).strip("\n").strip("\n")
    robofiyat[i] = (robofiyat[i].text).strip("\n").replace("\n","").strip("KDV Dahil").replace("KDV Dahil ","")
    try:
     roboeskifiyat[i] = (roboeskifiyat[i].text).strip("").replace("KDV Dahil"," ").strip() 
     robomEskiFiyat = roboeskifiyat[i].strip("\n").replace("","")
    except IndexError:
        robomEskiFiyat = ""
    roboliste.append([roboisim[i],robofiyat[i]])#,robomEskiFiyat

#roboblogn11
     
response = requests.get(url4)
html_icerigi = response.content
soup = BeautifulSoup(html_icerigi,"html.parser")

robon11eskifiyat = soup.find_all("div",{"class":"showcase-price-old"})
robon11fiyat = soup.find_all("span",{"class":"newPrice cPoint priceEventClick"})
robon11isim = soup.find_all("h3",{"class":"productName"})

robon11liste = list()

for i in range(len(robon11isim)):
 
    robomn11EskiFiyat = ""
    robon11isim[i] = (robon11isim[i].text).strip("\n").strip("\n").strip("               ")
    robon11fiyat[i] = (robon11fiyat[i].text).strip("\n").replace("\n","").strip("TL").replace("TL","")
    try:
     robon11eskifiyat[i] = (robon11eskifiyat[i].text).strip("").replace("KDV Dahil"," ").strip() 
     robomn11EskiFiyat = robon11eskifiyat[i].strip("\n").replace("","")
    except IndexError:
        robomn11EskiFiyat = ""
    robon11liste.append([robon11isim[i],robon11fiyat[i],dolarson2])#,robomEskiFiyat


df = pd.DataFrame(liste,columns = ["İsmi","Fiyat","dolar"])
dfr = pd.DataFrame(rliste,columns = ["İsmi","Fiyat"])
dfrobo = pd.DataFrame(roboliste,columns = ["İsmi","Fidyat"])
dfrobon11 = pd.DataFrame(robon11liste,columns = ["İsmi","Fidyat","dolar"])

#excel dosyası yazdırma
#wb = Workbook()
#ws = wb.active

#df.save("liste.xlsx")
#wb = Workbook()

#sayfa = wb.active

#planWorkbook = xlsxwriter.Workbook("C:\\Users\\3drea\\Desktop\\Siteveric-main\\deneme.xlsx")
#planSheet = planWorkbook.add_worksheet("fiyatlar")

#inputWorkbook = xlrd.open_workbook(df)

#df.to_excel("C:\\Users\\3drea\\Desktop\\Siteveric-main\\deneme1.xlsx", engine='xlsxwriter',sheet_name='3EKSEN'  )
#dfr.to_excel("C:\\Users\\3drea\\Desktop\\Siteveric-main\\deneme1.xlsx", engine='xlsxwriter',sheet_name='3dream'  ) 
with pd.ExcelWriter('C:\\Users\\3drea\\Desktop\\Siteveric-main\\deneme1.xlsx',engine='openpyxl',mode="a",if_sheet_exists="overlay") as writer:  
    df.to_excel(writer, sheet_name='3Eksen')
    dfr.to_excel(writer, sheet_name='3Dream') 
    dfrobo.to_excel(writer, sheet_name='Robobloq')  
    dfrobon11.to_excel(writer, sheet_name='Robobloq N11')


print(dfr)
print(df)
print(dfrobo)
print(dfrobon11)